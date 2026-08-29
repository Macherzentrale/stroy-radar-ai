import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_luxury_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мониторинг на Риска"
    ws.views.sheetView[0].showGridLines = True

    # Цветова палитра (Corporate Navy & Risk Palette)
    NAVY_DARK = "1E293B"      # Фон на главния хедър
    NAVY_LIGHT = "334155"     # Подзаглавия
    WHITE = "FFFFFF"
    BORDER_COLOR = "CBD5E1"
    
    # Цветове за статус на риска
    FILL_CRITICAL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Меко червено
    FONT_CRITICAL = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

    FILL_HIGH = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")     # Меко оранжево
    FONT_HIGH = Font(name="Segoe UI", size=10, bold=True, color="9A3412")

    FILL_MEDIUM = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")   # Меко жълто
    FONT_MEDIUM = Font(name="Segoe UI", size=10, bold=True, color="854D0E")

    FILL_LOW = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")      # Меко зелено
    FONT_LOW = Font(name="Segoe UI", size=10, bold=True, color="166534")

    # Тънки рамки
    thin_side = Side(border_style="thin", color=BORDER_COLOR)
    card_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1. Главен Header
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "  КОРПОРАТИВЕН РАДАР: РИСКОВИ СЪБИТИЯ И ЗАПОРИ В БЪЛГАРИЯ"
    title_cell.font = Font(name="Segoe UI", size=13, bold=True, color=WHITE)
    title_cell.fill = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    title_cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 40

    # 2. KPI Карти (Executive Summary Cards)
    kpis = [
        ("A3:B3", "A4:B4", "ОБЩО АНАЛИЗИРАНИ", "1 420 компании"),
        ("C3:D3", "C4:D4", "КРИТИЧЕН РИСК (ЗАПОРИ)", "14 субекта"),
        ("E3:F3", "E4:F4", "ОБЩА МАТЕРИАЛНА ЕКСПОЗИЦИЯ", "3 450 000 лв."),
        ("G3:H3", "G4:H4", "ДАТА НА ЕКСПОРТ", datetime.now().strftime("%d.%m.%Y | %H:%M"))
    ]

    for title_range, val_range, title, val in kpis:
        ws.merge_cells(title_range)
        ws.merge_cells(val_range)
        
        t_cell = ws[title_range.split(":")[0]]
        t_cell.value = title
        t_cell.font = Font(name="Segoe UI", size=8, bold=True, color="64748B")
        t_cell.alignment = Alignment(horizontal="center", vertical="center")
        t_cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        v_cell = ws[val_range.split(":")[0]]
        v_cell.value = val
        v_cell.font = Font(name="Segoe UI", size=11, bold=True, color=NAVY_DARK)
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        v_cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for row in range(3, 5):
        ws.row_dimensions[row].height = 20
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = card_border

    # 3. Данни за таблицата
    headers = [
        "ЕИК", "Име на дружеството", "Правна форма", 
        "Категория събитие", "Ниво на риск", "Детайли по изпълнителното дело", 
        "Дата на вписване", "Партида в ТР"
    ]
    
    records = [
        ("205849123", "ТЕХНО СОЛЮШЪНС ЕООД", "ЕООД", "Обезпечителна мярка / Запор", "CRITICAL", "Наложен запор от ЧСИ Георги Димитров (Рег. № 841) по ИД 20248410400120 за сума 45,200 лв.", "2026-08-28", "205849123"),
        ("131458992", "БАЛКАН ЛОГИСТИК ГРУП ООД", "ООД", "Фалит / Несъстоятелност", "HIGH", "Решение на СГС по т.д. 1420/2024. Назначен временен синдик.", "2026-08-27", "131458992"),
        ("201994821", "МЕДИКАЛ ДЕНТ БЪЛГАРИЯ АД", "АД", "Корпоративна промяна", "MEDIUM", "Заличаване на предишен управител. Встъпване на ново лице.", "2026-08-26", "201994821"),
        ("103859201", "АЛФА СТРОЙ ИНВЕСТ ЕООД", "ЕООД", "Осособен залог (ЦРОП)", "HIGH", "Учреден особен залог върху машини и оборудване към УниКредит.", "2026-08-25", "103859201"),
        ("204559182", "ДЕЛИВЪРИ ЕКСПРЕС ООД", "ООД", "Стандартна регистрация", "LOW", "Успешно подаден и вписан ГФО за предходната година.", "2026-08-25", "204559182"),
    ]

    # 4. Заглавия на колоните (Ред 6)
    table_header_row = 6
    ws.row_dimensions[table_header_row].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=table_header_row, column=col_idx, value=h)
        cell.font = Font(name="Segoe UI", size=9, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = card_border

    # 5. Запълване на данните (Редове 7+)
    start_row = 7
    for row_idx, r in enumerate(records, start_row):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(r, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = card_border
            cell.font = Font(name="Segoe UI", size=9)
            cell.alignment = Alignment(vertical="center")

            # Линк към регистъра
            if col_idx == 8:
                link_url = f"https://portal.registryagency.bg/CR/Reports/ActiveConditionTabResult?uic={val}"
                cell.value = "Преглед на живо ↗"
                cell.hyperlink = link_url
                cell.font = Font(name="Segoe UI", size=9, color="2563EB", underline="single")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Базово подравняване
            elif col_idx in [1, 3, 7]:
                cell.value = val
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = val

            # Цветово кодиране на риска
            if col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "CRITICAL":
                    cell.fill, cell.font = FILL_CRITICAL, FONT_CRITICAL
                elif val == "HIGH":
                    cell.fill, cell.font = FILL_HIGH, FONT_HIGH
                elif val == "MEDIUM":
                    cell.fill, cell.font = FILL_MEDIUM, FONT_MEDIUM
                elif val == "LOW":
                    cell.fill, cell.font = FILL_LOW, FONT_LOW

    # 6. Автоматично оразмеряване на колоните
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in ["A", "C", "E", "G", "H"]:
            ws.column_dimensions[col_letter].width = 16
        elif col_letter == "B":
            ws.column_dimensions[col_letter].width = 32
        elif col_letter == "D":
            ws.column_dimensions[col_letter].width = 28
        elif col_letter == "F":
            ws.column_dimensions[col_letter].width = 50

    # 7. Активиране на Auto-Filter
    last_row = start_row + len(records) - 1
    ws.auto_filter.ref = f"A{table_header_row}:H{last_row}"

    # 8. Запис на файла
    filename = f"Executive_Risk_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"\n[✓] ЛУКСОЗНИЯТ РЕПОРТ Е ГОТОВ: {filename}")

if __name__ == "__main__":
    generate_luxury_report()
